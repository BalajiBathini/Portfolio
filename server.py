import http.server
import socketserver
import json
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = 'form_data.xlsx'

# Ensure the Excel file exists
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name","contact", "Email","subject", "Message"])
    wb.save(EXCEL_FILE)

class MyHandler(http.server.SimpleHTTPRequestHandler):
    def do_POST(self):
        if self.path == '/submit':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            form_data = json.loads(post_data)
            
            # Append form data to the Excel file
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([form_data['name'], form_data['contact'], form_data['email'], form_data['subject'], form_data['message']])
            wb.save(EXCEL_FILE)
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'status': 'success'}).encode('utf-8'))
        else:
            self.send_error(404, 'File Not Found: %s' % self.path)

    def do_GET(self):
        if self.path == '/':
            self.path = 'index.html'
        return http.server.SimpleHTTPRequestHandler.do_GET(self)

PORT = 8000

with socketserver.TCPServer(("", PORT), MyHandler) as httpd:
    print(f"Serving at port {PORT}")
    httpd.serve_forever()
